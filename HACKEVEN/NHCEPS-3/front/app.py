import os
import cv2
import pytesseract
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, make_response, session
from werkzeug.utils import secure_filename
from google.oauth2 import service_account
from googleapiclient.discovery import build
import openai
from dotenv import load_dotenv
import openpyxl
from PIL import Image
import io

# Load environment variables
load_dotenv()

# Config
UPLOAD_FOLDER = os.path.join('static', 'uploads')
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# OpenAI configuration
openai.api_key = os.getenv('OPENAI_API_KEY')

# Google Sheets configuration
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')
SERVICE_ACCOUNT_FILE = 'billextracter-2d55c5313f14.json'

# App setup
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.secret_key = 'your-secret-key-here'  # Required for session

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Preprocessing function
def preprocess_image(path):
    image = cv2.imread(path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    inverted = cv2.bitwise_not(gray)
    resized = cv2.resize(inverted, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_LINEAR)
    thresh = cv2.threshold(resized, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    return thresh

# OCR function
def ocr_core(image_path):
    preprocessed = preprocess_image(image_path)
    text = pytesseract.image_to_string(preprocessed, config='--oem 3 --psm 6')
    return text

def get_google_sheets_service():
    try:
        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=credentials)
        return service
    except Exception as e:
        print(f"Error creating Google Sheets service: {str(e)}")
        return None

def process_with_ai(text):
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that extracts and categorizes bill information. Format the response as a structured list of items with categories, descriptions, and amounts. Return the data in a format that can be easily converted to a spreadsheet."},
                {"role": "user", "content": f"Please analyze this bill text and extract the items, categorize them, and provide their amounts. Format the response as a list of items with Category, Description, and Amount columns:\n\n{text}"}
            ]
        )
        
        ai_response = response.choices[0].message.content
        lines = ai_response.split('\n')
        structured_data = [['Category', 'Description', 'Amount']]  # Header row
        
        for line in lines:
            if ':' in line:
                parts = line.split(':')
                if len(parts) >= 2:
                    category = parts[0].strip()
                    details = parts[1].strip()
                    # Extract amount if present
                    if '$' in details:
                        amount = details.split('$')[-1].strip()
                        description = details.split('$')[0].strip()
                    else:
                        amount = '0'
                        description = details
                    structured_data.append([category, description, amount])
        
        return structured_data
    except Exception as e:
        print(f"AI processing error: {str(e)}")
        return None

def update_google_sheet(data):
    try:
        service = get_google_sheets_service()
        range_name = 'Sheet1!A1:C' + str(len(data))
        
        body = {
            'values': data
        }
        
        result = service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=range_name,
            valueInputOption='RAW',
            body=body
        ).execute()
        
        return result
    except Exception as e:
        print(f"Google Sheets update error: {str(e)}")
        return None

@app.route('/ai-process', methods=['POST'])
def ai_process():
    try:
        data = request.get_json()
        text = data.get('text')
        
        if not text:
            return jsonify({'error': 'No text provided'}), 400
        
        # Process with AI
        structured_data = process_with_ai(text)
        
        if not structured_data:
            return jsonify({'error': 'AI processing failed'}), 500
        
        # Update Google Sheet
        sheet_result = update_google_sheet(structured_data)
        
        if not sheet_result:
            return jsonify({'error': 'Google Sheets update failed'}), 500
        
        return jsonify({
            'success': True,
            'sortedData': structured_data,
            'sheetUrl': f'https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/', methods=['GET', 'POST'])
def index():
    ocr_text = None
    image_url = None
    
    if request.method == 'POST':
        if 'image' not in request.files:
            return redirect(request.url)
        file = request.files['image']
        if file.filename == '':
            return redirect(request.url)
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            text = ocr_core(file_path)
            image_url = url_for('static', filename=f'uploads/{filename}')

            return render_template('front.html', ocr_text=text, image_url=image_url)

    return render_template('front.html', ocr_text=ocr_text, image_url=image_url)

@app.route('/process', methods=['POST'])
def process_files():
    if 'image' not in request.files or 'excel' not in request.files:
        return 'Missing files', 400
    
    image_file = request.files['image']
    excel_file = request.files['excel']
    api_key = request.form.get('api_key')
    api_type = request.form.get('api_type', 'none')
    
    try:
        # Process image with OCR
        image = Image.open(image_file)
        extracted_text = pytesseract.image_to_string(image)
        
        # Process Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Update Excel with extracted text
        ws['A1'] = 'Extracted Text'
        ws['A2'] = extracted_text
        
        # Save the modified Excel file to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return both the text and Excel file
        response = make_response(send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='processed_bill.xlsx'
        ))
        
        # Add the extracted text to the response headers
        response.headers['X-Extracted-Text'] = extracted_text
        
        return response
        
    except Exception as e:
        print(f"Error processing files: {str(e)}")
        return str(e), 500

@app.route('/process-image', methods=['POST'])
def process_image():
    if 'image' not in request.files:
        return 'No image file provided', 400
    
    image_file = request.files['image']
    
    try:
        # Process image with OCR
        image = Image.open(image_file)
        extracted_text = pytesseract.image_to_string(image)
        
        if not extracted_text.strip():
            return 'No text could be extracted from the image', 400
        
        # Store the extracted text in session
        session['extracted_text'] = extracted_text
            
        return extracted_text
        
    except Exception as e:
        print(f"Error processing image: {str(e)}")
        return str(e), 500

@app.route('/process-excel', methods=['POST'])
def process_excel():
    if 'excel' not in request.files:
        return 'No Excel file provided', 400
    
    excel_file = request.files['excel']
    api_key = request.form.get('api_key')
    api_type = request.form.get('api_type', 'none')
    
    try:
        # Get the stored extracted text
        extracted_text = session.get('extracted_text')
        if not extracted_text:
            return 'No extracted text found. Please process an image first.', 400
        
        # Process Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Update Excel with extracted text
        # Split text into lines and write to Excel
        lines = extracted_text.split('\n')
        for i, line in enumerate(lines, start=1):
            ws[f'A{i}'] = line.strip()
        
        # Save the modified Excel file to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='processed_bill.xlsx'
        )
        
    except Exception as e:
        print(f"Error processing Excel: {str(e)}")
        return str(e), 500

if __name__ == '__main__':
    app.run(debug=True)